# Check_count_rows.py

import openpyxl

def check_count_rows(file_paths, res_path):
    count_rows_excel_files = 0
    numbers = []

    for file_path in file_paths:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        numbers.append(sheet.max_row)

    workbook = openpyxl.load_workbook(res_path)
    sheet = workbook.active
    count_rows_res_file = sheet.max_row

    for one in numbers:
        count_rows_excel_files += one

    print(f'Обработано {count_rows_res_file} строк, {len(numbers)} файлов')

    return count_rows_excel_files == count_rows_res_file