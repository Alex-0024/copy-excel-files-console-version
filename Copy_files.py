# Copy_files.py

import openpyxl
from Copy_cell_style import copy_cell_style

def copy_file(file_paths, res_path):
    # Создание новой рабочей книги для результата
    wb_result = openpyxl.Workbook()
    ws_result = wb_result.active

    # Переменная для отслеживания текущей строки в результате
    current_row = 1

    # Цикл по каждому исходному файлу
    for file_path in file_paths:
        # Загрузка текущего файла
        wb_source = openpyxl.load_workbook(file_path)
        ws_source = wb_source.active

        # Копирование данных из текущего файла
        for row in ws_source.iter_rows():
            for cell in row:
                new_cell = ws_result.cell(row=current_row, column=cell.column, value=cell.value)
                copy_cell_style(cell, new_cell)
            current_row += 1

    wb_result.save(res_path)