# Copy_row_column_sizes.py

import openpyxl

# Функция копирования ширины первой строки первого файла для всех строк итогового файла
def copy_row_column_sizes(begin_filename, res_filename):

    # Открыть исходный файл
    wb_source = openpyxl.load_workbook(begin_filename)
    ws_source = wb_source.active  # Предполагаем, что работаем с активным листом

    # Открыть целевой файл
    wb_target = openpyxl.load_workbook(res_filename)
    ws_target = wb_target.active  # Предполагаем, что работаем с активным листом

    # Взять высоту первой строки первого файла
    first_row_height = ws_source.row_dimensions[1].height

    # Установка высоты строк для всего целевого файла на основании первой строки первого файла из списка
    for ind in range (1, ws_target.max_row + 1):
        ws_target.row_dimensions[ind].height = first_row_height

    # Установка ширины столбцов целевого файла на основании первого файла из списка
    for col in ws_source.column_dimensions.keys():
        ws_target.column_dimensions[col].width = (
            ws_source.column_dimensions[col].width)

    # Сохранить изменения в целевом файле
    wb_target.save(res_filename)