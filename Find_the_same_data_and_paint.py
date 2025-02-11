# Find_the_same_data_and_paint

import openpyxl
from openpyxl.styles import PatternFill

# Функция копирования ширины первой строки первого файла для всех строк итогового файла
def find_data_and_paint(filename):

    # Открыть целевой файл
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Взять значение первой ячейки первой строки
    cell_value = sheet['A1'].value

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Определить строки с одинаковым значением в первой ячейке и закрасить
    for ind in range (2, sheet.max_row + 1):
        if sheet.cell(row=ind, column=1).value == cell_value:
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=ind, column=col).fill = yellow_fill

    # Сохранить изменения в целевом файле
    workbook.save(filename)